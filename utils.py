from openpyxl import load_workbook
from openpyxl.styles import Alignment
import re

# Função para ler os centros de custo (agora usando os dados das colunas B e C)
def ler_centros_custo(arquivo): 
    try:
        wb = load_workbook(arquivo, data_only=True)
        aba = wb.active
        dados_filtrados = []

        # Itera a partir da linha 2
        for row in aba.iter_rows(min_row=2, values_only=True):
            # Verifica se existem pelo menos 3 colunas (A, B, C)
            if len(row) >= 3:
                valorB = row[1]  # Coluna B
                valorC = row[2]  # Coluna C
                # Armazena os valores sem modificações
                dados_filtrados.append({"codigo": valorB, "descricao": valorC})
        wb.close()
        return dados_filtrados
    except Exception as e:
        return f"Erro ao ler os centros de custo: {e}"

def preencher_template(arquivo_template, dados_filtrados, grupo_aprovador):
    try:
        wb_template = load_workbook(arquivo_template)
        aba_template = wb_template.active

        if not dados_filtrados:
            raise Exception("Nenhum dado válido foi encontrado.")

        # Cria um objeto de alinhamento centralizado
        alinhamento_centro = Alignment(horizontal='left')

        for i, dado in enumerate(dados_filtrados, start=2):
            # Coluna 1: código (dados da coluna B da planilha base)
            cell = aba_template.cell(row=i, column=1, value=dado.get("codigo", ""))
            cell.alignment = alinhamento_centro

            # Coluna 2: descrição (dados da coluna C da planilha base)
            cell = aba_template.cell(row=i, column=2, value=dado.get("descricao", ""))
            cell.alignment = alinhamento_centro

            # Coluna 3: Ativo
            cell = aba_template.cell(row=i, column=3, value="Ativo")
            cell.alignment = alinhamento_centro

            # Preencher colunas de grupo, EXCLUINDO F (6) e L (12)
            # Colunas originais: [4, 6, 8, 10, 12, 14, 16, 18, 20, 22]
            # Ajustadas: [4, 8, 10, 14, 16, 18, 20, 22]
            for coluna in [4, 8, 10, 14, 16, 18, 20, 22]:
                cell = aba_template.cell(row=i, column=coluna, value="Grupo")
                cell.alignment = alinhamento_centro

            # Preencher colunas de aprovadores conforme solicitado.
            # Coluna E (5)
            cell = aba_template.cell(row=i, column=5, value=f"RC_NRM_1_{grupo_aprovador}")
            cell.alignment = alinhamento_centro
            # Coluna G (7) em branco
            cell = aba_template.cell(row=i, column=7, value="")
            cell.alignment = alinhamento_centro
            # Coluna I (9)
            cell = aba_template.cell(row=i, column=9, value=f"RC_ESP_1_{grupo_aprovador}")
            cell.alignment = alinhamento_centro
            # Coluna K (11)
            cell = aba_template.cell(row=i, column=11, value=f"RC_ESP_2_{grupo_aprovador}")
            cell.alignment = alinhamento_centro
            # Coluna M (13) em branco
            cell = aba_template.cell(row=i, column=13, value="")
            cell.alignment = alinhamento_centro
            # Coluna O (15)
            cell = aba_template.cell(row=i, column=15, value="PC_CDC_1")
            cell.alignment = alinhamento_centro
            # Coluna Q (17)
            cell = aba_template.cell(row=i, column=17, value="PC_GDC_1")
            cell.alignment = alinhamento_centro
            # Coluna S (19)
            cell = aba_template.cell(row=i, column=19, value=f"PC_DIR_{grupo_aprovador}")
            cell.alignment = alinhamento_centro
            # Coluna U (21) com valor fixo
            cell = aba_template.cell(row=i, column=21, value="PC_VPR_1")
            cell.alignment = alinhamento_centro
            # Coluna W (23)
            cell = aba_template.cell(row=i, column=23, value="PC_PRE_1")
            cell.alignment = alinhamento_centro

        # Salva as alterações no mesmo arquivo de template
        wb_template.save(arquivo_template)
        wb_template.close()
        return arquivo_template
    except Exception as e:
        raise Exception(f"Erro ao preencher o template: {e}")